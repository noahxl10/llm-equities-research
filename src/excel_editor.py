
import json
import os
import re
import threading
import traceback
from copy import copy
from dataclasses import dataclass
from datetime import datetime

import openpyxl
import pandas as pd
import yfinance as yf
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string

import src.config as config
from dba.data_models import Companies
from src.AZ import Azure
from src.utils import now
from ._types import ModelResponse, Company


@dataclass
class SuperExcel:
    """
    Base class providing utilities such as:
    - Alphabet indexing
    - Color mapping
    - Azure uploads
    """

    def __init__(self):
        mini_alphabet = [
            "A", "B", "C", "D", "E", "F", "G", "H", "I", "J",
            "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T",
            "U", "V", "W", "X", "Y", "Z"
        ]
        self.alphabet = mini_alphabet + [f"{a}{b}" for a in mini_alphabet for b in mini_alphabet]
        self.colors = {"red": "FF2D00", "green": "01DC13", "yellow": "F2FF19"}
        self.azure = Azure()

    def get_col_index(self, col):
        """Returns the 1-based index of a given column label."""
        return self.alphabet.index(col) + 1

    def upload_file(self, local_path, blob_name):
        """Uploads a local file to Azure Blob Storage."""
        try:
            self.azure.upload_blob(file_name=local_path, blob_name=blob_name, overwrite=True)
            self.azure.close()
            # os.remove(local_path)  # Uncomment if you want to delete the local file after upload
            return blob_name
        except Exception as e:
            print(e)


@dataclass
class FinancialModel(SuperExcel):
    """
    Builds and manipulates a financial model workbook (.xlsm) for a given company.
    Inserts new years/quarters, calculates projections, and updates formulas accordingly.
    """

    def __init__(self, company: Company):
        super().__init__()
        self.company = company
        self.ticker = company.ticker
        self.company_guid = company.guid

        self.base_file_path = f"{config.financial_model_file_name}.xlsm"
        self.export_path = f"{config.financial_model_file_name}_altered_{self.company.guid}.xlsm"
        self.blob_file_name = f"{config.financial_model_file_name_only}_altered.xlsm"

        self.tab = "Model"
        self.wb = openpyxl.load_workbook(self.base_file_path, read_only=False, keep_vba=True)
        self.ws = self.wb[self.tab]

        # Quarter map (e.g., "q1": ("mar", 1) means Q1 -> "mar" and numeric index 1)
        self.quarter_to_month_map = {
            "q1": ("mar", 1),
            "q2": ("jun", 2),
            "q3": ("sep", 3),
            "q4": ("dec", 4),
        }

        # Template constraints
        self.template_starting_column = "A"
        self.template_starting_year = 2007
        self.template_starting_year_row = 7
        self.template_ending_actual_year = 2022
        self.template_ending_actual_year_row = 22
        self.template_ending_yoy_year = 2022
        self.template_ending_yoy_year_row = 64
        self.ending_actual_year_row = None

        self.template_quarterly_indicator = "Quarterly"
        self.template_yoy_growth_indicator = "y/y growth"
        self.template_number_of_yoy_quarters = 19
        self.template_yoy_quarter_growth_indicator = "y/y growth q"
        self.template_starting_quarter = "q1mar18"
        self.template_ending_quarter = "q3sep23"
        self.template_number_of_actual_quarters = 23

        self.template_last_actual_year_row = 22
        self.template_last_actual_yoy_year_row = 64
        self.template_last_actual_quarter_row = 95
        self.template_last_actual_yoy_quarter_row = 127

        self.rows_added = 0
        self.added_columns = 0

        self.number_of_prediction_years = 6
        self.max_number_of_prediction_quarters = 8

        # Determine the most recent fiscal year and quarter
        self.recent_fiscal_year = self.get_most_recent_fiscal_year()
        self.recent_fiscal_quarter = 1  # If needed, uncomment below method to find actual quarter
        # self.recent_fiscal_quarter = self.get_most_recent_fiscal_quarter()

        # Quarter string e.g., "q1mar24"
        self.recent_fiscal_quarter_string = (
            f"q{self.recent_fiscal_quarter}"
            + self.quarter_to_month_map[f"q{self.recent_fiscal_quarter}"][0]
            + str(int(self.recent_fiscal_year) + 1)[2:]
        )

        # Holds references to financial docs (if needed)
        self.financial_documents = []

        # Insert new actual year rows if needed
        self.insert_new_actual_rows()
        self.update_projection_years()

        # Quarter rows
        self.starting_quarter_row = self.get_starting_quarter_row()
        self.ending_quarter_row = self.starting_quarter_row + self.template_number_of_actual_quarters
        self.insert_new_quarter_rows()
        self.update_projection_quarters()
        self.update_projection_sums()

        # Build year and quarter maps
        self.year_map = self.build_year_map()
        self.quarter_map = self.build_quarter_map()

        # Historical/prediction slicing for requests
        temp_years = [key for key in self.year_map]
        self.historical_years = temp_years[: temp_years.index(self.recent_fiscal_year) + 1]
        self.prediction_years = temp_years[temp_years.index(self.recent_fiscal_year) + 1 :]
        self.historical_years_for_requests = self.historical_years[-10:]

        temp_quarters = [key for key in self.quarter_map]
        self.historical_quarters = temp_quarters[: temp_quarters.index(self.recent_fiscal_quarter_string) + 1]
        self.historical_quarters_for_requests = self.historical_quarters[-10:]
        self.prediction_quarters = temp_quarters[temp_quarters.index(self.recent_fiscal_quarter_string) + 1 :]

        self.required_projection_years = self.recent_fiscal_year + self.number_of_prediction_years

    def get_projection_time_periods(self):
        """Collects all projection quarters and years as a combined list."""
        projection_time_periods = []
        projection_time_periods.extend(self.projection_quarters)
        projection_time_periods.extend([i + 1 for i in self.projection_years])
        return projection_time_periods

    def insert_new_actual_rows(self):
        """Inserts additional rows for new actual years if the company has more recent data."""
        dif = int(self.recent_fiscal_year) - int(self.template_ending_actual_year)
        if dif > 0:
            for i in range(dif):
                self.insert_row_below(
                    self.template_last_actual_year_row + i,
                    insert_type="year",
                    insert_value=self.template_ending_actual_year + i + 1,
                )
        self.ending_actual_year_row = self.template_last_actual_year_row + self.rows_added

    def update_projection_years(self):
        """Updates the next X years for projections in the sheet."""
        updated_projection_years = [
            self.template_ending_actual_year + i + 1 for i in range(self.number_of_prediction_years)
        ]
        self.projection_years = updated_projection_years

        for i, year in enumerate(updated_projection_years):
            row_offset = i + 1
            fy_row = self.ending_actual_year_row + row_offset
            actual_year = self.recent_fiscal_year + row_offset
            self.ws.cell(
                row=fy_row,
                column=column_index_from_string(self.template_starting_column),
            ).value = actual_year
        self.last_year = updated_projection_years[-1]

    def calculate_projection_rows(self, last_fiscal_year, last_fiscal_quarter):
        """
        Returns a list of future quarters to be projected based on last_fiscal_year/quarter.
        (Unused currently, but available for extended logic.)
        """
        projection_quarters = [
            "q1mar", "q2jun", "q3sep", "q4dec",
            "q1mar", "q2jun", "q3sep", "q4dec",
        ]
        projection_years = (
            [str(int(last_fiscal_year) + 1) for _ in range(4)]
            + [str(int(last_fiscal_year) + 2) for _ in range(4)]
        )
        combined = [
            f"{qt}{yr[2:]}" for qt, yr in zip(projection_quarters, projection_years)
        ]
        index = combined.index(last_fiscal_quarter)
        return combined[index + 1 :]

    def insert_new_quarter_rows(self):
        """Inserts additional rows if new actual quarters are more recent than template_ending_quarter."""
        dif = self.quarters_between(self.template_ending_quarter, self.recent_fiscal_quarter_string)
        self.template_last_actual_quarter_row += self.rows_added

        if dif > 0:
            for i in range(dif):
                self.insert_row_below(
                    self.template_last_actual_quarter_row + i,
                    insert_type="quarter",
                    insert_value=self.add_quarters(self.template_ending_quarter, i + 1),
                )
        self.ending_quarter_row = self.template_last_actual_quarter_row + dif

    def update_projection_quarters(self):
        """Updates quarters to be projected and places them in the sheet."""
        updated_projection_quarters = [
            self.add_quarters(self.recent_fiscal_quarter_string, i + 1)
            for i in range(self.max_number_of_prediction_quarters)
        ]
        self.projection_quarters = updated_projection_quarters

        for i, quarter in enumerate(updated_projection_quarters):
            row_offset = i + 1
            fy_row = self.ending_quarter_row + row_offset
            self.ws.cell(
                row=fy_row,
                column=column_index_from_string(self.template_starting_column),
            ).value = quarter
        self.last_quarter_row = fy_row

    def quarters_between(self, start, end):
        """Calculates how many quarters lie between two quarter strings (e.g., q1mar18, q3sep23)."""
        quarter_map = {"q1": 1, "q2": 2, "q3": 3, "q4": 4}

        start_quarter = quarter_map[start[:2].lower()]
        start_year = int(start[-2:])

        end_quarter = quarter_map[end[:2].lower()]
        end_year = int(end[-2:])

        total_start_quarters = start_year * 4 + start_quarter
        total_end_quarters = end_year * 4 + end_quarter

        return total_end_quarters - total_start_quarters

    def replace_formula_numbers(self, formula, start_number, end_number):
        """Replaces row numbers in SUM/AVERAGE formulas with new row indices."""
        pattern = re.compile(r"(SUM|AVERAGE)\([A-Z]+(\d+):[A-Z]+(\d+)\)")

        def replacement(match):
            function_name = match.group(1)
            # Extract the column letter from the matched formula
            part = match.group(0).split("(")[1].split(":")[0]
            column_letter = part[: -len(match.group(2))]
            return f"{function_name}({column_letter}{start_number}:{column_letter}{end_number})"

        return pattern.sub(replacement, formula)

    def get_cell_that_contains_string(self, search_string, starting_cell=None):
        """
        Searches for a cell whose value matches 'search_string'.
        Optionally start from a given row index (starting_cell).
        """
        if starting_cell is None:
            for i in range(150):
                cell_ref = f"{self.template_starting_column}{i+1}"
                cell_value = self.ws[cell_ref].value
                if str(cell_value) == search_string:
                    return i + 1
        else:
            for i in range(starting_cell, 200):
                cell_ref = f"{self.template_starting_column}{i+1}"
                cell_value = self.ws[cell_ref].value
                if cell_value and "=" in str(cell_value):
                    cell_name = cell_value.replace("=", "")
                    cell_value = self.ws[cell_name].value
                if str(cell_value) == search_string:
                    return i + 1

    def update_projection_sums(self):
        """Updates the SUM/AVERAGE formulas for the first two projected years in the sheet."""
        first_two_projection_year_rows = [
            self.get_cell_that_contains_string(search_string=f"{self.recent_fiscal_year+1}"),
            self.get_cell_that_contains_string(search_string=f"{self.recent_fiscal_year+2}")
        ]

        first_year = str(self.recent_fiscal_year + 1)[2:]
        second_year = str(self.recent_fiscal_year + 2)[2:]
        first_year_first_quarter = self.get_cell_that_contains_string(search_string=f"q1mar{first_year}")
        second_year_first_quarter = self.get_cell_that_contains_string(search_string=f"q1mar{second_year}")

        quarters = [
            [first_year_first_quarter, first_year_first_quarter + 3],
            [second_year_first_quarter, second_year_first_quarter + 3]
        ]
        row_check_range = [str(val) for val in range(90, 104)]

        for i, row in enumerate(first_two_projection_year_rows):
            for col in self.ws.iter_cols(min_row=row, max_row=row):
                for cell in col:
                    cell_val_str = str(cell.value)
                    if "SUM" in cell_val_str or "AVERAGE" in cell_val_str:
                        new_formula = self.replace_formula_numbers(
                            cell_val_str, quarters[i][0], quarters[i][1]
                        )
                        cell.value = new_formula
                    elif any(item in cell_val_str for item in row_check_range):
                        pattern = re.compile(r"[A-Za-z]+")
                        match = pattern.search(cell_val_str)
                        col_letter = match.group(0) if match else None
                        cell.value = f"={col_letter}{quarters[i][1]}"

    def build_year_map(self):
        """Builds a map of year -> row index for quick referencing."""
        year_map = {}
        for year in range(self.template_starting_year, self.last_year + 2):
            row_idx = self.template_starting_year_row + (year - self.template_starting_year)
            year_map[year] = row_idx
        return year_map

    def build_quarter_map(self):
        """Builds a map of quarter_string -> row index."""
        quarter_map = {}
        start_idx = self.ending_quarter_row - self.template_number_of_actual_quarters - 1
        for i in range(start_idx, self.last_quarter_row):
            row_quarter = self.add_quarters(self.template_starting_quarter, i - self.starting_quarter_row)
            quarter_map[row_quarter] = i
        return quarter_map

    def get_most_recent_fiscal_year(self) -> int:
        """
        Retrieves the most recent fiscal year for a given company's ticker via yfinance.
        Subtracts 1 year to ensure all data is 'closed out' (i.e., completed).
        """
        stock = yf.Ticker(self.company.ticker)
        financials = stock.financials
        financial_dates = financials.columns
        parsed_dates = [datetime.strptime(str(date), "%Y-%m-%d %H:%M:%S") for date in financial_dates]
        most_recent_date = max(parsed_dates)
        return most_recent_date.year - 1

    def get_most_recent_fiscal_quarter(self) -> int:
        """Optional method to retrieve the most recent quarter from yfinance data."""
        stock = yf.Ticker(self.company.ticker)
        quarterly_financials = stock.quarterly_financials
        quarterly_dates = quarterly_financials.columns
        parsed_dates = [datetime.strptime(str(date), "%Y-%m-%d %H:%M:%S") for date in quarterly_dates]
        most_recent_date = max(parsed_dates)
        fiscal_quarter = (most_recent_date.month - 1) // 3 + 1
        return fiscal_quarter

    def add_quarters(self, financial_quarter: str, num_quarters: int) -> str:
        """
        Given a quarter string like 'q1mar24', adds `num_quarters` and returns the new quarter string.
        """
        quarter_key = financial_quarter[:2].lower()  # e.g. 'q1'
        month_abbrev = financial_quarter[2:5]       # e.g. 'mar'
        year_val = int(financial_quarter[5:])

        quarter_index = self.quarter_to_month_map[quarter_key][1]
        total_quarters = quarter_index + num_quarters

        new_year = year_val + (total_quarters - 1) // 4
        new_quarter_index = (total_quarters - 1) % 4 + 1
        new_quarter = f"q{new_quarter_index}"
        new_month = self.quarter_to_month_map[new_quarter][0]
        return f"{new_quarter}{new_month}{new_year % 100:02d}"

    def calculate_quarters_since(self, current_quarter: int, current_year: int):
        """
        Calculates how many quarters have passed since self.starting_quarter up to the
        given current_quarter/current_year.
        """
        month_mapping = {
            "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
            "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12
        }
        quarter_start_month = {1: 1, 2: 4, 3: 7, 4: 10}

        month_abbr = self.starting_quarter[2:5]
        start_year_val = int(self.starting_quarter[5:])

        start_month = month_mapping[month_abbr]
        start_date = datetime(start_year_val + 2000, start_month, 1)

        current_start_month = quarter_start_month[current_quarter]
        end_date = datetime(current_year, current_start_month, 1)

        delta_months = (end_date.year - start_date.year) * 12 + (end_date.month - start_date.month)
        delta_quarters = delta_months // 3
        return delta_quarters

    def get_starting_quarter_row(self, search_string=None):
        """
        Retrieves the first row index after the 'Quarterly' label (template_quarterly_indicator).
        """
        if search_string is None:
            search_string = self.template_quarterly_indicator
        for i in range(150):
            cell_ref = f"{self.template_starting_column}{i+1}"
            cell_value = self.ws[cell_ref].value
            if cell_value == search_string:
                return i + 2

    def copy_cell_format(self, source_cell, target_cell):
        """Copies styling, formatting, number format, and alignment from source_cell to target_cell."""
        try:
            target_cell.font = copy(source_cell.font)
        except Exception:
            pass
        try:
            target_cell.border = copy(source_cell.border)
        except Exception:
            pass
        try:
            target_cell.fill = copy(source_cell.fill)
        except Exception:
            pass
        try:
            target_cell.number_format = source_cell.number_format
        except Exception:
            pass
        try:
            target_cell.protection = copy(source_cell.protection)
        except Exception:
            pass
        try:
            target_cell.alignment = copy(source_cell.alignment)
        except Exception:
            pass

    def update_formulas(self, sheet, insert_row):
        """Updates formulas in the sheet after inserting a new row (shifts references down by 1 row)."""
        for row in sheet.iter_rows(min_row=insert_row + 1, values_only=False):
            for cell in row:
                if cell.data_type == "f":  # 'f' indicates a formula
                    formula = cell.value
                    if formula:
                        new_formula = Translator(
                            formula, origin=cell.coordinate
                        ).translate_formula(f"{cell.column_letter}{cell.row + 1}")
                        cell.value = new_formula

    def insert_row_below(self, row_idx, insert_type, insert_value=None):
        """
        Inserts a row below row_idx and copies the row's data/formulas.
        If insert_type is 'year' or 'quarter', we also set the value in col=1 to insert_value.
        """
        self.ws.insert_rows(row_idx + 1)

        if insert_type in ["year", "quarter"]:
            self.rows_added += 1
            for col in range(1, self.ws.max_column + 1):
                source_cell = self.ws.cell(row=row_idx, column=col)
                target_cell = self.ws.cell(row=row_idx + 1, column=col, value=source_cell.value)
                # Update formulas in the new row
                if source_cell.value and source_cell.data_type == "f":
                    formula = source_cell.value
                    new_formula = Translator(
                        formula, origin=source_cell.coordinate
                    ).translate_formula(target_cell.coordinate)
                    target_cell.value = new_formula

                if col == 1 and insert_value is not None:
                    target_cell.value = insert_value

                self.copy_cell_format(source_cell, target_cell)

            self.update_formulas(self.ws, row_idx + 1)

    def shift_column_with_formulas(self, col, shift_by):
        """
        Shifts a single column (col) over by shift_by columns, including formula references.
        """
        for row in range(1, self.ws.max_row + 1):
            cell = self.ws.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                new_col_letter = get_column_letter(cell.column + shift_by)
                new_formula = Translator(
                    cell.value, origin=cell.coordinate
                ).translate_formula(f"{new_col_letter}{cell.row}")
                new_cell = self.ws.cell(row=row, column=cell.column + shift_by)
                new_cell.value = new_formula
            else:
                new_cell = self.ws.cell(row=row, column=cell.column + shift_by)
                new_cell.value = cell.value

            if cell.has_style:
                new_cell._style = cell._style
            if cell.hyperlink:
                new_cell.hyperlink = cell.hyperlink
            if cell.comment:
                new_cell.comment = cell.comment

            cell.value = None
            cell._style = None

    def shift_columns_with_formulas(self, start_col, end_col, shift_by):
        """
        Shifts all columns in [start_col, end_col] right by shift_by columns.
        Iterates from right to left to avoid overwriting.
        """
        for col in range(end_col, start_col - 1, -1):
            self.shift_column_with_formulas(col, shift_by)

    def set_cell_formatting(self, cell_ref, format_type):
        """Applies a number format (percent, currency, number) to a given cell reference."""
        format_dict = {
            "percent": "0.00%",
            "currency": "$#,##0.00",
            "number": "#,##0.0",
        }
        self.ws[cell_ref].number_format = format_dict[format_type]

    def center_cell(self, cell_ref):
        """Centers alignment of the cell."""
        self.ws[cell_ref].alignment = Alignment(horizontal="center")

    def set_col_width(self):
        """
        Sets column widths based on whether a reference cell (row=6) is populated or not.
        Column A is always set to a wider width.
        """
        empty_width = 2.00
        filled_width = 15.75
        check_row = 6

        for col in self.ws.columns:
            col_letter = col[0].column_letter
            cell_value = self.ws[f"{col_letter}{check_row}"].value
            if cell_value in ("", None):
                self.ws.column_dimensions[col_letter].width = empty_width
            else:
                self.ws.column_dimensions[col_letter].width = filled_width

        self.ws.column_dimensions["A"].width = filled_width

    def insert_segment_columns(self, segments):
        """
        Shifts columns to the right and inserts columns for each segment (3 columns per segment).
        For example: "segment rev", "pct of revenue", "growth rate".
        """
        first_col_seg_col_spacing = 2
        number_of_columns_per_segment = 3
        title_row = 6

        # Where to begin shifting
        start_column_index = column_index_from_string("C")
        end_column_index = column_index_from_string("IN")
        segment_shift = number_of_columns_per_segment * len(segments)

        self.shift_columns_with_formulas(start_column_index, end_column_index, shift_by=segment_shift)

        # Copy formatting from the first new column across all inserted columns
        start_col = column_index_from_string(self.template_starting_column) + 1
        end_col = number_of_columns_per_segment * len(segments) + start_col

        for row in range(1, self.ws.max_row + 1):
            source_cell = self.ws.cell(row=row, column=start_col)
            for col_idx in range(start_col + 1, end_col + 1):
                target_cell = self.ws.cell(row=row, column=col_idx)
                self.copy_cell_format(source_cell, target_cell)

        columns = [c for c in range(start_col + 1, end_col + 1)]
        seg_map = {}
        d_map = {}
        x = 0

        for seg in segments:
            seg_cols = columns[x : x + 3]
            seg_map[seg] = seg_cols
            x += 3
            d_map[seg_cols[0]] = "segment rev"
            d_map[seg_cols[1]] = "pct of revenue"
            d_map[seg_cols[2]] = "growth rate"

        self.segment_column_map = seg_map

        # Write headers
        for seg, cols in seg_map.items():
            for col_idx in cols:
                self.ws.column_dimensions[get_column_letter(col_idx)].width = 15
                self.ws.cell(row=title_row - 1, column=col_idx, value=seg)
                self.ws.cell(row=title_row, column=col_idx, value=d_map[col_idx])

        self.added_columns = len(segments) * 3

    def build_full_model_map(self):
        """
        (Optional/incomplete) placeholder to build a comprehensive mapping
        of all historical and projected data in the model.
        """
        pass

    def insert_responses(self, responses, segments=False):
        """
        Writes responses into the workbook based on year_map, quarter_map, and (optionally) segment columns.
        """
        if segments:
            # If responses is a dict keyed by segment -> {year -> {segment_percent_of_revenue, segment_growth_rate}}
            for segment_name in responses:
                cols = self.segment_column_map[segment_name]
                pct_col = cols[1]
                growth_col = cols[2]

                for year in responses[segment_name]:
                    row = None
                    # Find row from year_map
                    for key in self.year_map:
                        if str(key) == str(year).replace(" ", ""):
                            row = self.year_map[key]
                            break

                    # Insert the two columns
                    if row:
                        pct_val = responses[segment_name][year]["segment_percent_of_revenue"]
                        growth_val = responses[segment_name][year]["segment_growth_rate"]

                        self.ws[f"{get_column_letter(pct_col)}{row}"].value = pct_val
                        self.ws[f"{get_column_letter(growth_col)}{row}"].value = growth_val

        else:
            # responses is a list of measure dicts (including columns, time_period, etc.)
            for measure in responses:
                try:
                    if str(measure.get("is_yoy")) in ["1", "1.0"]:
                        # yoy data
                        if "q" in measure["time_period"]:
                            starting_row = 110
                        else:
                            starting_row = 55

                        row = self.get_cell_that_contains_string(measure["time_period"], starting_row)
                        col = get_column_letter(
                            column_index_from_string(measure["column"].replace("yoy", "")) + self.added_columns
                        )
                        measure["cell"] = f"{col}{row}"
                        self.ws[measure["cell"]].value = measure["amount"]

                    elif "," in measure["column"]:
                        # If measure["column"] has multiple columns
                        for col_part in measure["column"].split(","):
                            col_part = col_part.strip()
                            # Year map
                            for key in self.year_map:
                                if str(key) == str(measure["time_period"]).replace(" ", ""):
                                    measure["row"] = self.year_map[key]
                            # Quarter map
                            for q_key in self.quarter_map:
                                if q_key == measure["time_period"].replace(" ", ""):
                                    measure["row"] = self.quarter_map[q_key]

                            try:
                                col_letter = get_column_letter(
                                    column_index_from_string(col_part) + self.added_columns
                                )
                                measure["cell"] = f"{col_letter}{measure['row']}"
                                if self.ws[measure["cell"]].value is None:
                                    self.ws[measure["cell"]].value = measure["amount"]
                                elif "=" not in str(self.ws[measure["cell"]].value):
                                    self.ws[measure["cell"]].value = measure["amount"]
                            except Exception:
                                print(traceback.print_exc())
                    else:
                        # Single column
                        for key in self.year_map:
                            if str(key) == str(measure["time_period"]).replace(" ", ""):
                                measure["row"] = self.year_map[key]
                        for q_key in self.quarter_map:
                            if q_key == measure["time_period"].replace(" ", ""):
                                measure["row"] = self.quarter_map[q_key]

                        try:
                            col_letter = get_column_letter(
                                column_index_from_string(measure["column"]) + self.added_columns
                            )
                            measure["cell"] = f"{col_letter}{measure['row']}"
                            if self.ws[measure["cell"]].value is None:
                                self.ws[measure["cell"]].value = measure["amount"]
                            elif "=" not in str(self.ws[measure["cell"]].value):
                                self.ws[measure["cell"]].value = measure["amount"]
                        except Exception:
                            pass

                    # Apply formatting
                    self.set_cell_formatting(measure["cell"], measure["number_format_type"])
                    self.center_cell(measure["cell"])
                except Exception:
                    print(traceback.print_exc())

    def get_time_period_map(self):
        """
        Placeholder for a structured dictionary of historical vs. projection quarters/years.
        """
        return {
            "projection": {"quarter": "", "year": ""},
            "historical": {"quarter": "", "year": ""},
        }

    def projection_logic(self, response, column):
        """
        Example placeholder function that calculates a final value
        based on prior-year data or averages.
        """
        year_values = {"2018": 204.44, "2019": 163.44, "2020": 284.44, "2021": 361.44, "2022": 455.44, "2023": 404.44}
        year_ratios = {"2022": 0.33, "2023": 0.44}
        last_fiscal_year = "2023"

        last_year_value = year_values[last_fiscal_year]
        three_year_avg = (year_values["2021"] + year_values["2022"] + year_values["2023"]) / 3
        five_year_avg = (
            year_values["2019"] + year_values["2020"] + year_values["2021"] +
            year_values["2022"] + year_values["2023"]
        ) / 5
        last_year_ratio = year_ratios[last_fiscal_year]

        final_calc = ((last_year_value * last_year_ratio) + three_year_avg + five_year_avg) / 3
        return final_calc

    def build_financial_model(self, responses=None, segments_dict=None, segments=None):
        """
        Main entry point: Optionally inserts new segment columns, writes responses, and saves workbook.
        """
        if responses is None:
            responses = []
        if segments_dict is None:
            segments_dict = {}
        if segments is None:
            segments = []

        if segments:
            self.insert_segment_columns(segments)

        self.insert_responses(segments_dict, segments=True)
        self.insert_responses(responses, segments=False)
        self.set_col_width()
        self.export_workbook()

        blob_name = f'uploads/engines/{self.company_guid}/financial_model/{now("epoch")}/{self.blob_file_name}'
        # self.upload_file(local_path=self.export_path, blob_name=blob_name)
        return self.export_path, blob_name

    def export_workbook(self):
        """Saves and closes the workbook."""
        self.wb.save(self.export_path)
        self.wb.close()

    def get_cell_for_response(self, response: ModelResponse):
        """
        Example helper to return the cell reference for a given ModelResponse.
        (Unused if you're not populating data this way.)
        """
        response_year = response.year
        row = self.year_map[response_year]
        column = self.response_type_column_map[response.type]  # This dict is not defined in code
        return f"{column}{row}"


@dataclass
class OnePager(SuperExcel):
    """
    Builds and manipulates a 'one-pager' XLSM file for a given company.
    Inserts answers into a 'quality rating' table and colors cells accordingly.
    """

    def __init__(self, Company: Companies):
        super().__init__()
        self.ticker = Company.ticker
        self.company_guid = Company.guid
        self.base_file_path = f"{config.one_pager_file_name}.xlsm"
        self.export_path = f"{config.one_pager_file_name}_altered_{self.company_guid}.xlsm"
        self.blob_file_name = f"{config.one_pager_file_name_only}_altered.xlsm"

        self.sheet_name = "LLM"
        self.ticker_cell = "B5"
        self.thread_semaphore = threading.Semaphore(5)
        self.wb = openpyxl.load_workbook(self.base_file_path, read_only=False, keep_vba=True)
        self.ws = self.wb[self.sheet_name]

        self.table_map = {
            "issue_label": "O",
            "issue_name": "P",
            "y": "X",
            "n": "Y",
            "question": "Z",
            "y_n": "AH",
            "#": "AI",
            "answer": "EE",
            "y answer expression": "EM",
            "n answer expression": "EU",
        }

        self.qr_map = self.build_qr_map(self.ws)

    def build_qr_map(self, ws):
        """
        Example method to build a map of certain cells. 
        Currently references columns O, Y, etc. for 'Issue'.
        """
        cells = [f"{prefix}{i}" for prefix in ["O", "Y"] for i in range(2, 55)[7::5]]
        qr_map = {}
        for cell in cells:
            val = ws[cell].value
            if val not in ["Issue", None]:
                row = ws[cell].row
                col = ws[cell].column
                qr_map[val] = {
                    "sum_color": f"{self.alphabet[col - 1 + 9]}{row}",
                    "issues": [f"{self.alphabet[col - 1 + 1]}{r+1}" for r in range(row, row + 3)],
                    "color_issues": [f"{self.alphabet[col - 1 + 9]}{r+1}" for r in range(row, row + 3)],
                }
        return qr_map

    def fill_colors(self, result_number, cell):
        """Fills a cell with a color based on result_number > 0 (green), < 0 (red), = 0 (yellow)."""
        if int(result_number) > 0:
            cell_fill_color = self.colors["green"]
        elif int(result_number) < 0:
            cell_fill_color = self.colors["red"]
        else:
            cell_fill_color = self.colors["yellow"]

        self.ws[cell].fill = PatternFill(
            start_color=cell_fill_color,
            end_color=cell_fill_color,
            fill_type="solid"
        )

    def fill_quality_rating_table(self, df, company_segment_responses):
        """
        Main logic to fill the 'quality rating' table based on a DataFrame `df` and optional segment data.
        """
        col_label = self.table_map["issue_label"]
        row_range = range(59, 140)

        for row in row_range:
            subcat = self.ws[f"{col_label}{row}"].value

            if subcat not in ("Issue", None):
                if "=" in str(subcat):
                    subcat = self.ws[subcat.replace("=", "")].value

                issues = [
                    self.ws[f"{self.table_map['issue_name']}{row + i}"].value
                    for i in range(1, 4)
                ]
                sum_issues = 0

                for index, issue_name in enumerate(issues):
                    try:
                        y_n = df.loc[df["issue"] == issue_name, "y_n"].values[0]
                        # Mark Y or N
                        cell_y_n = f"{self.table_map['y_n']}{row + index + 1}"
                        self.ws[cell_y_n] = "Y" if y_n == 1 else "N"

                        # Write actual answer
                        answer_cell = f"{self.table_map['answer']}{row + index + 1}"
                        self.ws[answer_cell] = df.loc[df["issue"] == issue_name, "response"].values[0]

                        # Fill out "section 1" cells with Y or N expression
                        if y_n == 1:
                            y_expr_cell = f"{self.table_map['y answer expression']}{row + index + 1}"
                            self.ws[self.qr_map[subcat]["issues"][index]] = self.ws[y_expr_cell].value
                        else:
                            n_expr_cell = f"{self.table_map['n answer expression']}{row + index + 1}"
                            self.ws[self.qr_map[subcat]["issues"][index]] = self.ws[n_expr_cell].value

                        # Retrieve numeric result for coloring
                        try:
                            result_number = int(df.loc[df["issue"] == issue_name, "result_number"].values[0])
                        except ValueError:
                            result_number = 0
                        sum_issues += result_number

                    except Exception:
                        print(f"e1: {traceback.format_exc()}\n")

        # Place ticker
        self.ws[self.ticker_cell] = self.ticker

        # If there's segment data to fill, do it here
        if company_segment_responses:
            for response in company_segment_responses:
                try:
                    self.ws[response["cell"]] = response["response"]
                except Exception:
                    print(traceback.format_exc())

        self.wb.save(self.export_path)
        self.wb.close()

    def build_one_pager(self, df, segment_responses):
        """
        Main entry point for building the OnePager. 
        Fills the table with data from `df` and `segment_responses`, 
        then returns the path and blob name.
        """
        with self.thread_semaphore:
            self.fill_quality_rating_table(df, segment_responses)
            blob_name = f'uploads/engines/{self.company_guid}/one_pager/{now("epoch")}/{self.blob_file_name}'

        return self.export_path, blob_name

    def fill_one_company(self, ticker, local):
        """
        Example method for processing a single company's data.
        (Unused if you're not specifically handling file uploads and requests.)
        """
        df = pd.read_csv(os.path.join(self.UPLOAD_FOLDER, self.file))
        self.ticker = ticker
        # responses = send_requests(segment_requests)  # Not defined in snippet
        # Process responses
        # ...
        self.fill_quality_rating_table(ticker)


class EvidenceSynthesis(SuperExcel):
    """
    Builds and manipulates an 'evidence synthesis' workbook for a given company,
    injecting specific text data into certain cells.
    """

    def __init__(self, Company: Companies):
        super().__init__()
        self.company = Company
        self.base_file_path = f"{config.evidence_synthesis_file_name}.xlsx"
        self.export_path = f"{config.evidence_synthesis_file_name}_{self.company.guid}.xlsx"
        self.blob_file_name = f"{config.evidence_synthesis_file_name_only}_altered.xlsx"
        self.sheet_name = "Output"

        self.ticker_cell = "B2"
        self.investment_recommendation_cell = "C4"
        self.one_line_summary_cell = "C5"
        self.full_synthesis_cell = "C6"

    def fill_table(self, thesis_config):
        """Populates the evidence synthesis sheet with key text data."""
        self.ws[self.ticker_cell] = self.company.ticker
        self.ws[self.investment_recommendation_cell] = thesis_config["investment_recommendation"]
        self.ws[self.one_line_summary_cell] = thesis_config["one_line_summary"]
        self.ws[self.full_synthesis_cell] = thesis_config["full_synthesis"]

    def build(self, thesis_config):
        """Loads the template, fills the table, saves, and returns the local path + blob name."""
        self.wb = openpyxl.load_workbook(self.base_file_path, read_only=False)
        self.ws = self.wb[self.sheet_name]

        self.fill_table(thesis_config)
        self.wb.save(self.export_path)
        self.wb.close()

        blob_name = f'uploads/engines/{self.company.guid}/evidence_synthesis/{now("epoch")}/{self.blob_file_name}'
        return self.export_path, blob_name
